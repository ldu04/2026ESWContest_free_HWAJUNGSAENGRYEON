import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "재료비명세"

KFONT = "맑은 고딕"
thin = Side(style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_fill = PatternFill("solid", fgColor="D9E1F2")
tot_fill = PatternFill("solid", fgColor="F2F2F2")
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
right = Alignment(horizontal="right", vertical="center")

# Title
ws.merge_cells("A1:D1")
ws["A1"] = "재료비 명세 — 화중생련 「불사」 (16노드 기준)"
ws["A1"].font = Font(name=KFONT, size=13, bold=True)
ws["A1"].alignment = center
ws.row_dimensions[1].height = 26

# Header
headers = ["항목", "수량", "용도", "금액(원)"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=2, column=c, value=h)
    cell.font = Font(name=KFONT, size=10.5, bold=True)
    cell.fill = hdr_fill
    cell.alignment = center
    cell.border = border

items = [
    ("ESP32 개발보드 (WROOM-32D)", 18, "통신·감지·임종신호 담당 핵심 노드 (예비 포함)", 144000),
    ("DS18B20 온도센서 모듈", 17, "노드의 화재 접근(고온) 감지", 45900),
    ("라즈베리파이4 2GB 세트", 1, "데이터 수집·관제·대시보드 게이트웨이", 90000),
    ("라즈베리파이 전원 아답터", 1, "게이트웨이 전원 (5V 3A)", 5500),
    ("microSD 카드 64GB", 1, "게이트웨이 OS·데이터 저장", 25000),
    ("18650 배터리 셀", 4, "독립전원 노드 시연용", 34000),
    ("18650 배터리 홀더", 4, "배터리 장착", 2400),
    ("MT3608 승압 모듈", 4, "배터리 3.7V→5V 승압", 2800),
    ("18650 충전기", 1, "배터리 셀 충전", 6000),
    ("브레드보드 (830핀)", 2, "회로 구성·프로토타이핑", 2800),
    ("점퍼 케이블 (F-F 40P)", 2, "노드-센서 배선", 1700),
    ("WS2812 LED 모듈", 20, "노드 생사 상태 표시", 12000),
    ("마이크로 USB 케이블", 15, "노드 전원·펌웨어 업로드", 10500),
    ("납땜 세트", 1, "배터리 노드 배선 제작", 24000),
]

r = 3
for name, qty, use, amt in items:
    ws.cell(row=r, column=1, value=name).alignment = left
    ws.cell(row=r, column=2, value=qty).alignment = center
    ws.cell(row=r, column=3, value=use).alignment = left
    ac = ws.cell(row=r, column=4, value=amt); ac.alignment = right; ac.number_format = "#,##0"
    for c in range(1,5):
        cell = ws.cell(row=r, column=c)
        cell.border = border
        if cell.font.name != KFONT:
            cell.font = Font(name=KFONT, size=10)
    r += 1

first, last = 3, r-1
# 소계
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
ws.cell(row=r, column=1, value="소계 (상품금액, 부가세 별도)")
ws.cell(row=r, column=4, value=f"=SUM(D{first}:D{last})").number_format = "#,##0"
subtotal_row = r
r += 1
# VAT
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
ws.cell(row=r, column=1, value="부가세 (10%)")
ws.cell(row=r, column=4, value=f"=D{subtotal_row}*0.1").number_format = "#,##0"
vat_row = r
r += 1
# 배송비
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
ws.cell(row=r, column=1, value="배송비")
ws.cell(row=r, column=4, value=2500).number_format = "#,##0"
ship_row = r
r += 1
# 합계
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
ws.cell(row=r, column=1, value="합계 (실결제 예정액)")
ws.cell(row=r, column=4, value=f"=D{subtotal_row}+D{vat_row}+D{ship_row}").number_format = "#,##0"
total_row = r

for rr in range(subtotal_row, total_row+1):
    for c in range(1,5):
        cell = ws.cell(row=rr, column=c)
        cell.border = border
        cell.fill = tot_fill
        bold = (rr == total_row)
        cell.font = Font(name=KFONT, size=(11 if bold else 10), bold=bold)
        if c in (1,):
            cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c == 4:
            cell.alignment = right

# note
r = total_row + 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
note = ws.cell(row=r, column=1, value="※ 디바이스마트 실구매가 기준(2026), 국내 배송. 노드 16개 + 예비 포함. 데모 구조물(폼보드)·공용 충전기는 별도.")
note.font = Font(name=KFONT, size=9, italic=True, color="666666")
note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 8
ws.column_dimensions["C"].width = 38
ws.column_dimensions["D"].width = 14

wb.save("재료비명세_불사_16노드.xlsx")
print("saved")
